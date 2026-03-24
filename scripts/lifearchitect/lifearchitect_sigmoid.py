#!/usr/bin/env python3
"""
docs/aime_sigmoid Task 2 — Sigmoid frontier fits for LifeArchitect.xlsx.

Reads `tables/LifeArchitect.xlsx` (sheet: "Models Table"), filters to rows where:
  - Parameters (B) is present and numeric
  - Tokens trained (B) is present and numeric
  - Tokens trained (B) cell is NOT italic (important: italic indicates an estimate)

Then fits a high-quantile sigmoid frontier for each target:
  - MMLU
  - MMLU-Pro
  - GPQA

Methodology and plotting style match:
  `scripts/epoch_ai/fit_benchmarks_runs_sigmoid_frontiers.py`

Outputs:
  outputs/lifearchitect/sigmoid/no_split/{points,curves,plots}/
"""

from __future__ import annotations

import argparse
import math
import os
import sys
from typing import Dict, Iterable, Optional, Sequence, Tuple

import numpy as np
from openpyxl import load_workbook

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from skill_frontier.core.fit_imports import import_fit_sigmoid_frontier_extended  # type: ignore
try:
    from skill_frontier.io.output_paths import sanitize_task_name
except Exception:

    def sanitize_task_name(task: str) -> str:  # type: ignore[misc]
        name = task.replace(" ", "_")
        for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            name = name.replace(ch, "_")
        return name


fit_sigmoid_frontier = import_fit_sigmoid_frontier_extended()

try:
    from skill_frontier.plotting.axis_formatting import apply_pretraining_compute_tick_multiplier
    from skill_frontier.plotting.configs import frontier_1d as frontier_1d_cfg
    from skill_frontier.plotting.configs import mpl_rc as mpl_rc_cfg
    from skill_frontier.plotting.labels import PRETRAINING_COMPUTE_FLOPS_LABEL
except Exception:  # pragma: no cover
    apply_pretraining_compute_tick_multiplier = None  # type: ignore

    class _Frontier1DCfg:  # type: ignore
        FIGSIZE = (7.0, 4.6)
        SCATTER_SIZE = 12
        SCATTER_ALPHA = 0.20
        SCATTER_LINEWIDTHS = 0.0
        CURVE_LINEWIDTH = 2.4
        CURVE_ALPHA = 0.95
        LABEL_FONTSIZE_X = 18
        LABEL_FONTSIZE_Y = 18
        TITLE_FONTSIZE = 18
        TICK_LABELSIZE = 15
        TICK_LENGTH = 4.0
        TICK_WIDTH = 0.9
        TICK_DIRECTION = "out"
        GRID_MAJOR_LINESTYLE = "-"
        GRID_MAJOR_LINEWIDTH = 0.9
        GRID_MAJOR_COLOR = "#d0d0d0"
        GRID_MAJOR_ALPHA = 0.7
        GRID_MINOR_LINESTYLE = ":"
        GRID_MINOR_LINEWIDTH = 0.6
        GRID_MINOR_COLOR = "#e0e0e0"
        GRID_MINOR_ALPHA = 0.7
        SPINE_LINEWIDTH = 1.1
        SPINE_COLOR = "#555555"
        LEGEND_FONTSIZE = 15
        LEGEND_FRAMEALPHA = 0.15
        LEGEND_LOC = "best"

    frontier_1d_cfg = _Frontier1DCfg()  # type: ignore

    class _MplRcCfg:  # type: ignore
        FONT_FAMILY = "serif"

    mpl_rc_cfg = _MplRcCfg()  # type: ignore
    PRETRAINING_COMPUTE_FLOPS_LABEL = "Pretraining Compute (FLOPs)"


X_TICK_MULTIPLIER: float = 1e21
TARGET_FIGSIZE: tuple[float, float] = (0.75 * (2069.0 / 300.0), 1344.0 / 300.0)
HEADER_ROW = 2
DATA_START_ROW = 3


_COMPUTE_OVERRIDES: Dict[str, Dict[str, str]] = {
    # Nemotron Elastic reports the *elastic/distillation training* token budget (110B),
    # not the teacher's pretraining budget. For compute-scaling plots we want the base
    # model's pretraining compute as the x-axis.
    "Nemotron Elastic": {"base_model": "Nemotron Nano 2"},
}


def _canon(s: object) -> str:
    import re

    txt = str(s or "").strip().lower().replace("\n", " ")
    txt = re.sub(r"\s+", " ", txt)
    return re.sub(r"[^a-z0-9]+", "", txt)


def _to_float(x: object) -> float:
    try:
        if x is None:
            return float("nan")
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "" or s.lower() in {"nan", "none", "na"}:
            return float("nan")
        return float(s)
    except Exception:
        return float("nan")


def _apply_frontier_style(ax, *, xlabel: str, ylabel: str, title: str) -> None:
    ax.set_xscale("log")
    ax.set_xlabel(xlabel, fontweight="bold", fontsize=frontier_1d_cfg.LABEL_FONTSIZE_X)
    if apply_pretraining_compute_tick_multiplier is not None and xlabel == PRETRAINING_COMPUTE_FLOPS_LABEL:
        apply_pretraining_compute_tick_multiplier(ax)
    ax.set_ylabel(ylabel, fontweight="bold", fontsize=frontier_1d_cfg.LABEL_FONTSIZE_Y)
    ax.set_title(title, fontweight="bold", fontsize=frontier_1d_cfg.TITLE_FONTSIZE)
    ax.tick_params(
        axis="both",
        labelsize=frontier_1d_cfg.TICK_LABELSIZE,
        length=frontier_1d_cfg.TICK_LENGTH,
        width=frontier_1d_cfg.TICK_WIDTH,
        direction=frontier_1d_cfg.TICK_DIRECTION,
    )
    try:
        import matplotlib.ticker as mticker  # type: ignore

        ax.yaxis.set_minor_locator(mticker.AutoMinorLocator(2))
        if ax.get_xscale() == "log":
            ax.xaxis.set_minor_locator(mticker.LogLocator(base=10.0, subs=(0.2, 0.4, 0.6, 0.8)))
    except Exception:
        pass

    ax.yaxis.grid(
        True,
        which="major",
        linestyle=frontier_1d_cfg.GRID_MAJOR_LINESTYLE,
        linewidth=frontier_1d_cfg.GRID_MAJOR_LINEWIDTH,
        color=frontier_1d_cfg.GRID_MAJOR_COLOR,
        alpha=frontier_1d_cfg.GRID_MAJOR_ALPHA,
    )
    ax.yaxis.grid(
        True,
        which="minor",
        linestyle=frontier_1d_cfg.GRID_MINOR_LINESTYLE,
        linewidth=frontier_1d_cfg.GRID_MINOR_LINEWIDTH,
        color=frontier_1d_cfg.GRID_MINOR_COLOR,
        alpha=frontier_1d_cfg.GRID_MINOR_ALPHA,
    )
    ax.xaxis.grid(
        True,
        which="major",
        linestyle=frontier_1d_cfg.GRID_MAJOR_LINESTYLE,
        linewidth=frontier_1d_cfg.GRID_MAJOR_LINEWIDTH,
        color=frontier_1d_cfg.GRID_MAJOR_COLOR,
        alpha=frontier_1d_cfg.GRID_MAJOR_ALPHA,
    )
    ax.xaxis.grid(
        True,
        which="minor",
        linestyle=frontier_1d_cfg.GRID_MINOR_LINESTYLE,
        linewidth=frontier_1d_cfg.GRID_MINOR_LINEWIDTH,
        color=frontier_1d_cfg.GRID_MINOR_COLOR,
        alpha=frontier_1d_cfg.GRID_MINOR_ALPHA,
    )
    for spine in ax.spines.values():
        spine.set_linewidth(frontier_1d_cfg.SPINE_LINEWIDTH)
        spine.set_color(frontier_1d_cfg.SPINE_COLOR)


def _resolve_columns(ws) -> Dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v is None:
            continue
        k = _canon(v)
        if k:
            headers[k] = c

    def pick(*aliases: str) -> int:
        for a in aliases:
            if a in headers:
                return headers[a]
        raise KeyError(f"Missing required column. Tried: {aliases}. Present keys: {sorted(headers)[:40]}...")

    return {
        "model": pick("model"),
        "params_b": pick("parametersb", "paramsb"),
        "tokens_b": pick("tokenstrainedb", "tokenstrainingb"),
        "mmlu": pick("mmlu"),
        "mmlu_pro": pick("mmlupro", "mmlupro"),
        "gpqa": pick("gpqa"),
    }


def _iter_rows(ws, cols: Dict[str, int]) -> Iterable[Tuple[str, float, float, bool, float, float, float]]:
    for r in range(DATA_START_ROW, ws.max_row + 1):
        model = ws.cell(row=r, column=cols["model"]).value
        if model is None or str(model).strip() == "":
            continue
        model_s = str(model).strip()

        params_b = _to_float(ws.cell(row=r, column=cols["params_b"]).value)
        tok_cell = ws.cell(row=r, column=cols["tokens_b"])
        tokens_b = _to_float(tok_cell.value)
        tokens_italic = bool(tok_cell.font.italic)

        mmlu = _to_float(ws.cell(row=r, column=cols["mmlu"]).value)
        mmlu_pro = _to_float(ws.cell(row=r, column=cols["mmlu_pro"]).value)
        gpqa = _to_float(ws.cell(row=r, column=cols["gpqa"]).value)
        yield (model_s, params_b, tokens_b, tokens_italic, mmlu, mmlu_pro, gpqa)


def _fit_one(
    *,
    out_dir: str,
    task_name: str,
    tau: float,
    kappa: float,
    lambda_b: float,
    min_models: int,
    models: Sequence[str],
    compute_scaled: np.ndarray,
    scores_frac: np.ndarray,
) -> None:
    if compute_scaled.size < int(min_models):
        print(f"[lifearchitect] skip task={task_name}: n={compute_scaled.size} < {min_models}")
        return

    safe_task = sanitize_task_name(task_name)
    out_points = os.path.join(out_dir, "sigmoid", "no_split", "points")
    out_curves = os.path.join(out_dir, "sigmoid", "no_split", "curves")
    out_plots = os.path.join(out_dir, "sigmoid", "no_split", "plots")
    for d in (out_points, out_curves, out_plots):
        os.makedirs(d, exist_ok=True)

    points_path = os.path.join(out_points, f"{safe_task}.csv")
    with open(points_path, "w", encoding="utf-8") as f:
        f.write("model,training_compute_flop,log10_compute,score\n")
        for m, x_scaled, y in zip(models, compute_scaled, scores_frac):
            c_abs = float(x_scaled) * float(X_TICK_MULTIPLIER)
            z = math.log10(c_abs) if c_abs > 0 else float("nan")
            f.write(f"{m},{c_abs:.12g},{z:.12g},{float(y):.6g}\n")

    xs_curve_scaled, y_curve = fit_sigmoid_frontier(
        compute_scaled,
        scores_frac,
        tau=float(tau),
        use_log10_x=True,
        fit_mode="quantile_per_point",
        kappa_final=float(kappa),
        lambda_b=float(lambda_b),
    )
    if xs_curve_scaled.size == 0:
        print(f"[lifearchitect] WARN: fit failed for task={task_name} (n={compute_scaled.size})")
        return

    curve_path = os.path.join(out_curves, f"{safe_task}.csv")
    with open(curve_path, "w", encoding="utf-8") as f:
        f.write("training_compute_flop,y_hat,task\n")
        for x_scaled, yv in zip(xs_curve_scaled, y_curve):
            c_abs = float(x_scaled) * float(X_TICK_MULTIPLIER)
            f.write(f"{c_abs:.12g},{float(yv):.12g},{task_name}\n")

    import matplotlib as mpl  # type: ignore
    import matplotlib.pyplot as plt  # type: ignore

    mpl.rcParams["font.family"] = mpl_rc_cfg.FONT_FAMILY
    fig, ax = plt.subplots(figsize=TARGET_FIGSIZE)
    ax.scatter(
        compute_scaled,
        scores_frac,
        s=float(frontier_1d_cfg.SCATTER_SIZE) * 3.0,
        alpha=0.6,
        color="black",
        marker="X",
        label="points",
        linewidths=frontier_1d_cfg.SCATTER_LINEWIDTHS,
        rasterized=True,
    )
    ax.plot(
        xs_curve_scaled,
        y_curve,
        color="firebrick",
        linewidth=frontier_1d_cfg.CURVE_LINEWIDTH,
        label=f"Smooth τ={float(tau):.2f}",
        alpha=frontier_1d_cfg.CURVE_ALPHA,
    )

    _apply_frontier_style(
        ax,
        xlabel=PRETRAINING_COMPUTE_FLOPS_LABEL,
        ylabel="Best score",
        title="",
    )
    y_min = float(np.nanmin([np.nanmin(scores_frac), np.nanmin(y_curve)]))
    y_max = float(np.nanmax([np.nanmax(scores_frac), np.nanmax(y_curve)]))
    if not np.isfinite(y_min) or not np.isfinite(y_max) or y_min == y_max:
        y_min, y_max = 0.0, 1.0
    pad = 0.02 * max(1e-6, (y_max - y_min))
    ax.set_ylim(y_min - pad, y_max + pad)
    ax.legend(
        loc=frontier_1d_cfg.LEGEND_LOC,
        fontsize=frontier_1d_cfg.LEGEND_FONTSIZE,
        frameon=True,
        framealpha=frontier_1d_cfg.LEGEND_FRAMEALPHA,
    )
    fig.tight_layout()
    out_base = os.path.join(out_plots, safe_task)
    fig.savefig(out_base + ".png", dpi=300)
    fig.savefig(out_base + ".pdf", dpi=300)
    plt.close(fig)


def main(argv: Optional[Sequence[str]] = None) -> None:
    ap = argparse.ArgumentParser(description="Fit LifeArchitect sigmoid frontiers (MMLU/MMLU-Pro/GPQA)")
    ap.add_argument("--xlsx", default=os.path.join("tables", "LifeArchitect.xlsx"))
    ap.add_argument("--sheet", default="Models Table")
    ap.add_argument("--out_dir", default=os.path.join("outputs", "lifearchitect"))
    ap.add_argument("--tau", type=float, default=0.98)
    ap.add_argument("--kappa", type=float, default=50.0)
    ap.add_argument("--lambda_b", type=float, default=1e-4)
    ap.add_argument("--min_models", type=int, default=20)
    args = ap.parse_args(argv)

    wb = load_workbook(str(args.xlsx), data_only=True)
    ws = wb[str(args.sheet)] if str(args.sheet) in wb.sheetnames else wb.active
    cols = _resolve_columns(ws)

    rows = list(_iter_rows(ws, cols))
    # Base filter: params + tokens present and tokens not italic.
    base = [
        r
        for r in rows
        if (not r[3])
        and np.isfinite(r[1])
        and r[1] > 0
        and np.isfinite(r[2])
        and r[2] > 0
    ]

    # Precompute compute for every model so we can override derived models to use their base model compute.
    compute_by_model: Dict[str, float] = {}
    for (model, params_b, tokens_b, _italic, _mmlu, _mmlu_pro, _gpqa) in base:
        tokens_t = float(tokens_b) / 1000.0
        compute_scaled = float(6.0 * float(params_b) * tokens_t)
        if np.isfinite(compute_scaled) and compute_scaled > 0:
            compute_by_model[model] = compute_scaled

    def _maybe_override_compute(model: str, compute_scaled: float) -> float:
        override = _COMPUTE_OVERRIDES.get(model)
        if not override:
            return compute_scaled
        base_model = override.get("base_model")
        if not base_model:
            return compute_scaled
        base_compute = compute_by_model.get(base_model)
        if base_compute is None or not (np.isfinite(base_compute) and base_compute > 0):
            print(
                f"[lifearchitect] WARN: compute override requested for {model!r} "
                f"but base model {base_model!r} missing/invalid; using original compute."
            )
            return compute_scaled
        if not np.isfinite(compute_scaled) or compute_scaled <= 0:
            return base_compute
        if base_compute <= compute_scaled:
            return base_compute
        # If the base compute is larger, it indicates the original tokens are a partial/retraining budget.
        print(
            f"[lifearchitect] compute override: {model!r} uses base_model={base_model!r} "
            f"(compute_scaled {compute_scaled:.6g} -> {base_compute:.6g})"
        )
        return base_compute

    def build_xy(score_idx: int) -> Tuple[Sequence[str], np.ndarray, np.ndarray]:
        models = []
        xs = []
        ys = []
        for (model, params_b, tokens_b, _italic, mmlu, mmlu_pro, gpqa) in base:
            score = (mmlu, mmlu_pro, gpqa)[score_idx]
            if not np.isfinite(score):
                continue
            tokens_t = float(tokens_b) / 1000.0
            compute_scaled = float(6.0 * float(params_b) * tokens_t)
            compute_scaled = _maybe_override_compute(model, compute_scaled)
            if not (np.isfinite(compute_scaled) and compute_scaled > 0):
                continue
            y = float(score) / 100.0
            if not np.isfinite(y):
                continue
            y = float(np.clip(y, 0.0, 1.0))
            models.append(model)
            xs.append(compute_scaled)
            ys.append(y)
        return models, np.array(xs, float), np.array(ys, float)

    task_specs = [
        ("MMLU", 0),
        ("MMLU-Pro", 1),
        ("GPQA", 2),
    ]
    os.makedirs(str(args.out_dir), exist_ok=True)
    for task_name, idx in task_specs:
        models, x, y = build_xy(idx)
        _fit_one(
            out_dir=str(args.out_dir),
            task_name=task_name,
            tau=float(args.tau),
            kappa=float(args.kappa),
            lambda_b=float(args.lambda_b),
            min_models=int(args.min_models),
            models=models,
            compute_scaled=x,
            scores_frac=y,
        )

    print(f"[lifearchitect] outputs in: {args.out_dir}")


if __name__ == "__main__":  # pragma: no cover
    main()
